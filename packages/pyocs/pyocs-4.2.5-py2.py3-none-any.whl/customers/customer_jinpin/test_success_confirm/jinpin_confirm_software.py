from openpyxl import load_workbook
import os
from pyocs import pyocs_software
from customers.customer_jinpin.test_success_confirm import ocs_sendemail

# 创建对象
confirm_task = pyocs_software.PyocsSoftware()

send_email = ocs_sendemail.my_osc_sendemail()



# 打开Excel文件读取数据
dirPath = os.getcwd()
Workbook = load_workbook('客户回复确认表格.xlsx')
# 获取行列等单元值
sheet = Workbook.active
max_rows_new = sheet.max_row

to_engyneer = ['linxiangna@cvte.com','chenfan3714@cvte.com','chenchaoxiong@cvte.com','liyangqiu@cvte.com'
                ]

# main
row_index = 1
ocs_number = 000000
soft_version = str()
order_info = str()
fangan_info = str()
order_soft_success_list = list()
order_soft_fail_list = list()
chenchaoxiong_unpass_order = list()
linxiangna_unpass_order = list()
chengfan_unpass_order = list()
zhanghonghai_unpass_order = list()

chenchaoxiong_upload_ocs_sucess_order = list()
linxiangna_upload_ocs_sucess_order = list()
chengfan_upload_ocs_sucess_order = list()
zhanghonghai_upload_ocs_sucess_order = list()
upload_soft_to_ocs_success_list = list()
text_Enter = '\n'




while row_index <= max_rows_new:
    order_info = sheet.cell(row=row_index, column=1)
    soft_version = sheet.cell(row=row_index, column=3)
    fangan_info = sheet.cell(row=row_index, column=2)
    if None != order_info.value and str(soft_version.value).rstrip(" "):
        tmp_success_list = [order_info.value, soft_version.value,fangan_info.value]
        order_soft_success_list.append(tmp_success_list)
    else:#找出评审单号为空的确认不通过的订单
        if(None != fangan_info.value):
            tmp_fail_list = [order_info.value, soft_version.value, fangan_info.value]
            #print("评审单号为空：",row_index)
            order_soft_fail_list.append(tmp_fail_list)

    row_index += 1
print(max_rows_new)
print(order_soft_success_list)
print("failed")
print(order_soft_fail_list)

for key in order_soft_success_list:
    ret = confirm_task.set_software_confirm(key[1], key[0])
    if ret:
        print(key[0] + " -- " + key[1] + " confirm success")
        ocs_number=confirm_task.get_ocs_number_from_sw(key[1], key[0])
        #print(ocs_number)
        key.append(ocs_number)
        #print(key)
        upload_soft_to_ocs_success_list.append(key)
    else:#找出评审单位不为空的确认不通过的订单

        tmp_fail_list = key
        order_soft_fail_list.append(tmp_fail_list)

for key in upload_soft_to_ocs_success_list:
    #根据方案找order_soft_fail_list的对应工程师
    engineer_name=ocs_sendemail.my_osc_sendemail.get_engineer_name(key[2])
    #print(engineer_name)
    #print(key)
    tmp_fail_list=key
    if("chenchaoxiong@cvte.com"==engineer_name):
        chenchaoxiong_upload_ocs_sucess_order.append(tmp_fail_list)
    if ("linxiangna@cvte.com" == engineer_name):
        linxiangna_upload_ocs_sucess_order.append(tmp_fail_list)
    if ("chenfan3714@cvte.com" == engineer_name):
        chengfan_upload_ocs_sucess_order.append(tmp_fail_list)
    if ("zhanghonghai@cvte.com" == engineer_name):
        zhanghonghai_upload_ocs_sucess_order.append(tmp_fail_list)

for key in order_soft_fail_list:
    #根据方案找order_soft_fail_list的对应工程师
    engineer_name=ocs_sendemail.my_osc_sendemail.get_engineer_name(key[2])
    #print(engineer_name)
    #print(key)
    tmp_fail_list=key
    if("chenchaoxiong@cvte.com"==engineer_name):
        chenchaoxiong_unpass_order.append(tmp_fail_list)
    if ("linxiangna@cvte.com" == engineer_name):
        linxiangna_unpass_order.append(tmp_fail_list)
    if ("chenfan3714@cvte.com" == engineer_name):
        chengfan_unpass_order.append(tmp_fail_list)
    if ("zhanghonghai@cvte.com" == engineer_name):
        zhanghonghai_unpass_order.append(tmp_fail_list)

if(0 != len(order_soft_fail_list)):
    new_str_chenfang = send_email.failed_list_to_str(chengfan_unpass_order)
    new_str_chenchaoxiong = send_email.failed_list_to_str(chenchaoxiong_unpass_order)
    new_str_xiangna = send_email.failed_list_to_str(linxiangna_unpass_order)
    new_str_honghai = send_email.failed_list_to_str(zhanghonghai_unpass_order)
    upload_ocs_sucess_str_chenfang = send_email.failed_list_to_str(chengfan_upload_ocs_sucess_order)
    upload_ocs_sucess_str_chenchaoxiong = send_email.failed_list_to_str(chenchaoxiong_upload_ocs_sucess_order)
    upload_ocs_sucess_str_xiangna = send_email.failed_list_to_str(linxiangna_upload_ocs_sucess_order)
    upload_ocs_sucess_str_honghai = send_email.failed_list_to_str(zhanghonghai_upload_ocs_sucess_order)
    confirm_success_text = "亲爱的工程师，您好，以下订单自动确认通过"
    upload_ocs_sucess_text_chenfang = text_Enter + "陈凡:" + text_Enter + upload_ocs_sucess_str_chenfang
    upload_ocs_sucess_text_chenchaoxiong = text_Enter + "潮雄:" + text_Enter + upload_ocs_sucess_str_chenchaoxiong
    upload_ocs_sucess_text_xiangna = text_Enter + "祥纳:" + text_Enter + upload_ocs_sucess_str_xiangna
    upload_ocs_sucess_text_honghai = text_Enter + "宏海:" + text_Enter + upload_ocs_sucess_str_honghai
    confirm_success_list = confirm_success_text+upload_ocs_sucess_text_chenfang + upload_ocs_sucess_text_chenchaoxiong + upload_ocs_sucess_text_xiangna + upload_ocs_sucess_text_honghai
    #confirm_success_list = confirm_success_text+text_Enter+send_email.failed_list_to_str(upload_soft_to_ocs_success_list)+ text_Enter
    text = "亲爱的工程师，您好，以下订单无法自动确认通过，请及时关注"
    text_chenfang = text_Enter + "陈凡:" + text_Enter + new_str_chenfang
    text_chenchaoxiong = text_Enter + "潮雄:" + text_Enter + new_str_chenchaoxiong
    text_xiangna = text_Enter + "祥纳:" + text_Enter + new_str_xiangna
    text_honghai = text_Enter + "宏海:" + text_Enter + new_str_honghai
    whole_text = confirm_success_list+text + text_chenfang + text_chenchaoxiong + text_xiangna + text_honghai
    """
    def get_whole_text():
        i=0
        text = "亲爱的工程师，您好，以下订单无法自动确认通过，请及时关注"
        str_4_engineer_name=["陈凡：","王航：","祥纳：","宏海："]
        for i in range(0,3)
            str=send_email.failed_list_to_str()
            whole_text = text_Enter + str_4_engineer_name[i] + text_Enter+text +str
        return whole_text
    print(whole_text)
    """
    send_email.send_email_to_engineer(to_engyneer,whole_text)