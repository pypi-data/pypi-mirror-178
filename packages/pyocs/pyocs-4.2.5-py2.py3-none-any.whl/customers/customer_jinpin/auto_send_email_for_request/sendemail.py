#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
from email import encoders
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import parseaddr, formataddr
import smtplib
import os
from pyocs import pyocs_login

def _format_addr(s):
    name, addr = parseaddr(s)
    return formataddr((Header(name, 'utf-8').encode(), addr))



# Workbook属于创建才用
# 打开Excel文件读取数据
dirPath = os.getcwd()
readbook = load_workbook('追需求表格2.xlsx')
account = pyocs_login.PyocsLogin().get_account_from_json_file()
from_addr = str(account['Username'])+"@cvte.com"
password = account['Password']

to_engyneer = ['linxiangna@cvte.com', 'chenfan3714@cvte.com', 'chenchaoxiong@cvte.com',
               'liyangqiu@cvte.com', 'duanguoqing@cvte.com',
               'guanbao.zhang@jpe.cc', 'yunfeng.chang@jpe.cc', 'jiayuan.chen@jpe.cc', 'xiaocong.cen@jpe.cc',
               'xuebo.luo@jpe.cc', 'kaijian.lin@jpe.cc','dajun.li@jpe.cc',
               'jia.sun@jpe.cc'
               ]
# to_engyneer=sned_to_target.split(",")
smtp_server = 'smtp.exmail.qq.com'
text_str1 = 'Dear  各位金品电子工程师：'
text_str2 = '\n'
text_str3 = '     您们好！请提供如上软件需求及参考软件版本，如果需求已提供请忽略此邮件，谢谢！'
subject_str1 = '【新订单处理 】'
subject_str2 = '祥鑫（BTY）   '
subject_str3 = '机型:E55EK2100'
subject_str4 = 'TP.HV553.PC821(4K*2K)'
subject_str5 = '评审单号:'
subject_str6 = '01-201903-105'
subject_str7 = '孟加拉'
subject_str8 = '     '
kongge = '  '


str_fangan="     方案对应工程师如下："
str_chenchaoxiong="     陈潮雄 56,3700,69,6306,5507,358,5522"
str_linxiangna="     林祥纳 59,3553,512,638,338,5510,8503,960,ATM30"
str_chenfan="     陈凡 506,310,320,510,530,553,3663,3458"
str_zhanghonghai="     张宏海 3686"
str_fangan_beizhu=text_str2+str_fangan+text_str2+str_chenchaoxiong+text_str2+str_linxiangna+text_str2+str_chenfan+text_str2+str_zhanghonghai
text_str = text_str1 + text_str2 + text_str3+str_fangan_beizhu
# 打开excel
#readbook = load_workbook(dir_str1)
# 获取读入的文件的sheet
sheet = readbook.active
max_rows_new = sheet.max_row
row_index = 1

while row_index < max_rows_new:
    b = sheet.cell(row=row_index, column=1)
    # if b.value:
    if None != b.value:
        subject_str2 = sheet.cell(row=row_index, column=6).value
        subject_str7 = sheet.cell(row=row_index, column=7).value
        subject_str3 = sheet.cell(row=row_index, column=9).value
        subject_str4 = sheet.cell(row=row_index, column=13).value
        subject_str6 = b.value
        # 不接受有这些关键项为空的情况
        subject_str = subject_str1 + subject_str2 + kongge + subject_str7 + subject_str8 + subject_str3 +\
                      subject_str8 + subject_str4 + subject_str8 + subject_str5 + subject_str6
        #msg = MIMEMultipart('alternative')
        msg = MIMEText(text_str, 'plain', 'utf-8')  # text_str发送正文，固定语句+确认不通过的订单信息
        #msgTEXT2 = MIMEText(str_fangan_beizhu, 'plain', 'utf-8')
        #msg.attach(msgTEXT1)
        #msg.attach(msgTEXT2)
        #msg_co = MIMEText(text_str, 'plain', 'utf-8')
        msg['From'] = _format_addr('CVTE_林祥纳<%s>' % from_addr)
        msg['To'] = ';'.join(to_engyneer)  # 以逗号形式连接起来 #_format_addr('管理员 <%s>' % to_addr)#可以采用把列表分开来的方法，一个一个发
        msg['Subject'] = Header(subject_str, 'utf-8').encode()

        server = smtplib.SMTP(smtp_server, 25)
        server.set_debuglevel(1)
        server.login(from_addr, password)
        server.sendmail(from_addr, to_engyneer, msg.as_string())
    row_index += 1
server.quit()
