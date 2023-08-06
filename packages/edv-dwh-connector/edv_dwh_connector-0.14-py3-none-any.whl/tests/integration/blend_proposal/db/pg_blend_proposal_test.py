"""
Test case for blend proposals from PostgreSQL.
.. since: 0.5
"""

# -*- coding: utf-8 -*-
# Copyright (c) 2022 Endeavour Mining
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to read
# the Software only. Permissions is hereby NOT GRANTED to use, copy, modify,
# merge, publish, distribute, sublicense, and/or sell copies of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NON-INFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

from datetime import date

import numpy as np  # type: ignore
import openpyxl  # type: ignore
import sqlalchemy  # type: ignore
from hamcrest import assert_that, raises, is_, calling, equal_to,\
    has_length, contains_inanyorder

from edv_dwh_connector.blend_proposal.db.pg_blend_proposal \
    import PgBlendProposal, PgBlendProposals
from edv_dwh_connector.blend_proposal.excel.blend_proposal_start_cell\
    import BlendProposalStartCell
from edv_dwh_connector.blend_proposal.excel.excel_blend_proposal\
    import ExcelBlendProposal
from edv_dwh_connector.exceptions import ValueNotFoundError
from tests.adapted_postgresql import AdaptedPostreSQL, PgDwhForTests

BLEND_FILE_NAME = "3.Daily_Blend_Template_process_October.xlsx"


def test_adds_blend_with_no_sequence() -> None:
    """
    Test adds blend with no sequence.
    """

    with AdaptedPostreSQL() as postgres:
        dwh = PgDwhForTests(
            sqlalchemy.create_engine(postgres.get_connection_url())
        )
        blends = PgBlendProposals(dwh)
        assert_that(
            calling(blends.add).with_args(
                PgBlendProposal(date(2022, 1, 1), dwh)
            ),
            raises(
                ValueError, "Blend proposal to add has no sequence"
            ),
            "Should not be able to add a blend proposal without sequence"
        )


def test_adds_a_blend_proposal(resource_path_root) -> None:
    """
    Test that its adds well blend.
    :param resource_path_root: Resource path
    """

    with AdaptedPostreSQL() as postgres:
        dwh = PgDwhForTests(
            sqlalchemy.create_engine(postgres.get_connection_url())
        )
        blends = PgBlendProposals(dwh)
        sheet = openpyxl.load_workbook(
            resource_path_root / BLEND_FILE_NAME
        ).active
        blend = blends.add(
            ExcelBlendProposal(
                sheet=sheet,
                bp_start_cell=BlendProposalStartCell(
                    sheet=sheet, day=date(2022, 10, 8)
                )
            )
        )
        assert_that(
            blends.items(
                date(2022, 10, 1), date(2022, 10, 31)
            ), has_length(2),
            "Should count blend proposals"
        )
        assert_that(
            blends.last().date(),
            equal_to(date(2022, 10, 8)),
            "Should get last blend proposal"
        )
        assert_that(
            blends.has_at(date(2022, 10, 8)), is_(True),
            "Should contains a blend proposal by its date"
        )
        assert_that(
            blend.sequences().items(), has_length(2),
            "Should count blend sequences"
        )
        assert_that(
            [seq.name() for seq in blend.sequences().items()],
            contains_inanyorder(
                "ON GOING", "IF LEP_HG_FR_LCU  IS  FINISHED"
            ),
            "Blend sequences names should match"
        )


def test_tries_to_add_a_wrong_blend(resource_path_root) -> None:
    """
    Test that it rollbacks blend.
    :param resource_path_root: Resource path
    """

    with AdaptedPostreSQL() as postgres:
        dwh = PgDwhForTests(
            sqlalchemy.create_engine(postgres.get_connection_url())
        )
        blends = PgBlendProposals(dwh)
        sheet = openpyxl.load_workbook(
            resource_path_root / "blend_proposal_with_formula.xlsx"
        ).active
        dte = date(2022, 11, 9)
        try:
            blends.add(
                ExcelBlendProposal(
                    sheet=sheet,
                    bp_start_cell=BlendProposalStartCell(
                        sheet=sheet, day=dte
                    )
                )
            )
        # pylint: disable=bare-except
        except:  # noqa: E722
            assert_that(
                blends.has_at(dte), equal_to(False),
                "Should not contain wrong blend"
            )


def test_gets_a_blend_proposal() -> None:
    """
    Test that its gets well blend.
    """

    with AdaptedPostreSQL() as postgres:
        dwh = PgDwhForTests(
            sqlalchemy.create_engine(postgres.get_connection_url())
        )
        blends = PgBlendProposals(dwh)
        assert_that(
            blends.get_at(date(2022, 10, 7)).date(),
            equal_to(date(2022, 10, 7)),
            "Should get a blend proposal by its date"
        )
        assert_that(
            blends.has_at(date(2022, 10, 7)), is_(True),
            "Should contains a blend proposal by its date"
        )
        assert_that(
            calling(blends.get_at).with_args(date(2022, 10, 6)),
            raises(
                ValueNotFoundError,
                "Blend proposal not found at date 2022-10-06"
            ),
            "Should not get a blend proposal at date"
        )


def test_tries_to_add_a_blend_with_undefined_values(
    resource_path_root
) -> None:
    """
    Test that it is able to add a blend with undefined values.
    :param resource_path_root: Resource path
    """

    with AdaptedPostreSQL() as postgres:
        dwh = PgDwhForTests(
            sqlalchemy.create_engine(postgres.get_connection_url())
        )
        blends = PgBlendProposals(dwh)
        sheet = openpyxl.load_workbook(
            resource_path_root / "blend_proposal_with_formula.xlsx",
            data_only=True
        ).active
        dte = date(2022, 11, 9)
        blend = blends.add(
            ExcelBlendProposal(
                sheet=sheet,
                bp_start_cell=BlendProposalStartCell(
                    sheet=sheet, day=dte
                )
            )
        )
        seq = blend.sequences().items()[2]
        assert_that(
            np.isnan(seq.recovery_blend_estimated()), is_(True),
            "Blend recovery estimated should not be a number"
        )
        assert_that(
            np.isnan(seq.materials().items()[2].available_tons()), is_(True),
            "Available tons should not be a number"
        )
