"""
    This submodule contain the class and functions used to read and write the DeepDetection Excel documents.
"""
import logging
import numpy as np
import numpy.typing as npt
from typing import List
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)

# Typing
NDArrayUint32 = npt.NDArray[np.uint32]


class ExcelOperate:
    """
        This class is used to operate Excel docs.
    """

    def __init__(self, filepath: str):
        """
        Initialize Excel_Operate object
        :param filepath: Absolute path to excel document
        """
        try:
            self.path = filepath
            # data_only ignores loading formulas and instead loads only the resulting values.
            self.workbook = load_workbook(filename=filepath, data_only=True)
            self.sheet_names = self.workbook.sheetnames
        except FileNotFoundError:
            logger.error(f"Path error, file not found: '{filepath}'")
            raise FileNotFoundError

    def save(self) -> int:
        """
        Save work book
        :return: If correct return positive number
        """
        try:
            self.workbook.save(self.path)
            return 0
        except Exception as e:
            logger.error(e)
            return -1

    def range_val_to_matrix(self, ranges: List[str], sheet_name: str) -> NDArrayUint32:
        """
        Used for read chip register table
        :param ranges: Excel cells to get, Ex: ["B2:AE20", ]
        :param sheet_name: Excel sheet name
        :return: Chip register matrix
        """
        sheet = self.workbook[sheet_name]

        tables = []
        for _range in ranges:
            # Read table values
            values = sheet[_range]
            container = np.zeros((np.shape(values)))  # Create matrix with equal size as values read
            row = 0
            for row_vals in values:
                column = 0
                for val in row_vals:
                    container[row][column] = val.value
                    column += 1
                row += 1
            tables.append(container)

        return np.array(tables, dtype=np.uint32)

    def get_n_sheet_matrix(self, ranges: List[str], sheet_pos: str) -> NDArrayUint32:
        """
        Used for read pixel register table
        :param ranges: Excel cells to get, Ex: ["B2:AE20", "B20:AE20"]
        :param sheet_pos: (Matrix) The positions in sheet where the info is placed.
        :return: (Array) Pixel register matrix.
        """
        mx = []
        if sheet_pos == "all":
            i = 0
            for sheet_name in self.sheet_names:  # Getting all sheets except the first one
                if not i == 0:
                    mx.append(self.range_val_to_matrix(ranges, sheet_name))
                i += 1

        else:
            sheet_name = self.sheet_names[sheet_pos]
            mx.append(self.range_val_to_matrix(ranges, sheet_name))

        return np.array(mx)

    def matrix_to_range_val(self, ranges: List[str], sheet_name: str, matrix: NDArrayUint32, bool_flag):
        """
        Used for write chip register and pixel register table
        :param ranges: Excel cells to get, Ex: ["B2:AE20", ]
        :param sheet_name: Excel sheet name
        :param matrix: Chip Register matrix to write
        :param bool_flag: True if the data write is BOOLEAN data
        """
        sheet = self.workbook[sheet_name]
        chip_iterate = 0
        for _range in ranges:
            row_iterate = 0
            for row in sheet.iter_rows(min_row=_range[0], max_row=_range[1], min_col=_range[2], max_col=_range[3]):
                column_iterate = 0
                if len(matrix) == 1:  # The code is analyzing the chip register matrix
                    if row_iterate == 11 or row_iterate == 12 or row_iterate == 17 or row_iterate == 18:
                        bool_flag = True
                    else:
                        bool_flag = False

                for val in row:
                    if bool_flag:
                        val.value = bool(matrix[chip_iterate][row_iterate][column_iterate])
                    else:
                        val.value = matrix[chip_iterate][row_iterate][column_iterate]
                    column_iterate += 1
                row_iterate += 1
            chip_iterate += 1

    def send_pixel_reg_matrix(self, ranges: List[str], matrix: NDArrayUint32):
        """
        Used for write pixel register table
        :param ranges: Excel cells to get, Ex: ["B2:AE20", ]
        :param matrix: Chip Register matrix to write
        """
        _bool = False
        i = 0
        for sheet_name in self.sheet_names[1:]:  # Getting all sheets except the first one.
            # Add all the sheet positions where dtypes are boolean
            if i == 0 or i == 1 or i == 2 \
                    or i == 4 or i == 5 or i == 6 \
                    or i == 8 or i == 9 or i == 10 \
                    or i == 12 or i == 13 or i == 14 \
                    or i == 16 or i == 17 or i == 18 \
                    or i == 20 or i == 21 or i == 22 \
                    or i == 24 or i == 25 or i == 26 \
                    or i == 28 or i == 29 or i == 30 \
                    or i == 32 or i == 33 or i == 34 \
                    or i == 35 or i == 36 or i == 41 or i == 42 or i == 43:
                _bool = True
            else:
                _bool = False

            self.matrix_to_range_val(ranges, sheet_name, matrix[i], _bool)
            i += 1


def get_linda_matrix(absolute_path: str) -> (NDArrayUint32, NDArrayUint32):
    """
    Get data from LINDA excel
    :param absolute_path: Absolute path to excel document
    :return: Chip register matrix and pixel register matrix
    """
    try:
        workbook = ExcelOperate(absolute_path)
    except InvalidFileException:
        logger.error("openpyxl does not support  file format, please check you can open it with Excel first."
                     " Supported formats are: .xlsx,.xlsm,.xltx,.xltm")
        raise FileNotFoundError

    # Getting chip register matrix
    chip_reg_range = ["B2:AE20", ]
    sheet_name = "ChipReg"

    chip_reg = workbook.range_val_to_matrix(chip_reg_range, sheet_name)
    logger.info(np.shape(chip_reg))

    # Getting pixel register matrix
    pixel_reg_ranges = ["C6:V13", "Y6:AR13", "C16:V23", "Y16:AR23", "C26:V33", "Y26:AR33",
                        "C36:V43", "Y36:AR43", "C46:V53", "Y46:AR53", "C56:V63", "Y56:AR63",
                        "C66:V73", "Y66:AR73", "C76:V83", "Y76:AR83", "C86:V93", "Y86:AR93",
                        "C96:V103", "Y96:AR103", "C106:V113", "Y106:AR113", "C116:V123", "Y116:AR123",
                        "C126:V133", "Y126:AR133", "C136:V143", "Y136:AR143", "C146:V153", "Y146:AR153"]
    sheet_pos_pixel_reg = "all"

    pixel_reg = workbook.get_n_sheet_matrix(pixel_reg_ranges, sheet_pos_pixel_reg)
    logger.info(np.shape(pixel_reg))

    return chip_reg, pixel_reg


def write_linda_matrix(absolute_path: str, chip_reg: NDArrayUint32, pixel_reg: NDArrayUint32) -> None:
    """
    Write data to LINDA excel
    :param absolute_path: Absolute path to excel document
    :param chip_reg: Chip register data
    :param pixel_reg: Pixel register data
    """
    workbook = ExcelOperate(absolute_path)

    # Writing chip register matrix
    chip_reg_range = [[2, 20, 2, 31], ]
    sheet_name = "ChipReg"
    workbook.matrix_to_range_val(chip_reg_range, sheet_name, chip_reg, False)  # I will use for C

    # Writing pixel register matrix
    pixel_reg_ranges = [[6, 13, 3, 22], [6, 13, 25, 44], [16, 23, 3, 22], [16, 23, 25, 44], [26, 33, 3, 22],
                        [26, 33, 25, 44]
        , [36, 43, 3, 22], [36, 43, 25, 44], [46, 53, 3, 22], [46, 53, 25, 44], [56, 63, 3, 22], [56, 63, 25, 44]
        , [66, 73, 3, 22], [66, 73, 25, 44], [76, 83, 3, 22], [76, 83, 25, 44], [86, 93, 3, 22], [86, 93, 25, 44]
        , [96, 103, 3, 22], [96, 103, 25, 44], [106, 113, 3, 22], [106, 113, 25, 44]
        , [116, 123, 3, 22], [116, 123, 25, 44], [126, 133, 3, 22], [126, 133, 25, 44], [136, 143, 3, 22]
        , [136, 143, 25, 44], [146, 153, 3, 22], [146, 153, 25, 44]]

    workbook.send_pixel_reg_matrix(pixel_reg_ranges, pixel_reg)

    error = workbook.save()
    if error < 0:
        logger.error("Can't save file, check if xlsx is opened!")
        raise OSError

