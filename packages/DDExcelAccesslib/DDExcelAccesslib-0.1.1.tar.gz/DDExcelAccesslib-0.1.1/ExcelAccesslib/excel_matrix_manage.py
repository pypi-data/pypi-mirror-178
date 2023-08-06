"""
    This submodule contain the class and functions used to read and write the DeepDetection Excel documents.
"""
import logging
import numpy as np
import numpy.typing as npt
from typing import List
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Typing
NDArrayUint32 = npt.NDArray[np.uint32]


class ExcelOperate:
    """
        This class is used to operate Excel docs.
    """

    def __init__(self, filepath: str):
        """
        Initialize Excel_Operate object
        :param filepath: Absolute path to excel document
        """
        try:
            self.path = filepath
            # data_only ignores loading formulas and instead loads only the resulting values.
            self.workbook = load_workbook(filename=filepath, data_only=True)
            self.sheet_names = self.workbook.sheetnames
        except FileNotFoundError:
            logger.error(f"Path error, file not found: '{filepath}'")
            raise FileNotFoundError

    def save(self) -> int:
        """
        Save work book
        :return: If correct return positive number
        """
        try:
            self.workbook.save(self.path)
            return 0
        except Exception as e:
            logger.error(e)
            return -1

    def range_val_to_matrix(self, ranges: List[str], sheet_name: str) -> NDArrayUint32:
        """
        Used for read chip register table
        :param ranges: Excel cells to get, Ex: ["B2:AE20", ]
        :param sheet_name: Excel sheet name
        :return: Chip register matrix
        """
        sheet = self.workbook[sheet_name]

        tables = []
        for _range in ranges:
            # Read table values
            values = sheet[_range]
            container = np.zeros((np.shape(values)))  # Create matrix with equal size as values read
            row = 0
            for row_vals in values:
                column = 0
                for val in row_vals:
                    container[row][column] = val.value
                    column += 1
                row += 1
            tables.append(container)

        return np.array(tables, dtype=np.uint32)

    def get_n_sheet_matrix(self, ranges: List[str], sheet_pos: str) -> NDArrayUint32:
        """
        Used for read pixel register table
        :param ranges: Excel cells to get, Ex: ["B2:AE20", "B20:AE20"]
        :param sheet_pos: (Matrix) The positions in sheet where the info is placed.
        :return: (Array) Pixel register matrix.
        """
        mx = []
        if sheet_pos == "all":
            i = 0
            for sheet_name in self.sheet_names:  # Getting all sheets except the first one
                if not i == 0:
                    mx.append(self.range_val_to_matrix(ranges, sheet_name))
                i += 1

        else:
            sheet_name = self.sheet_names[sheet_pos]
            mx.append(self.range_val_to_matrix(ranges, sheet_name))

        return np.array(mx)

    def matrix_to_range_val(self, ranges: List[str], sheet_name: str, matrix: NDArrayUint32, bool_flag):
        """
        Used for write chip register and pixel register table
        :param ranges: Excel cells to get, Ex: ["B2:AE20", ]
        :param sheet_name: Excel sheet name
        :param matrix: Chip Register matrix to write
        :param bool_flag: True if the data write is BOOLEAN data
        """
        sheet = self.workbook[sheet_name]
        chip_iterate = 0
        for _range in ranges:
            row_iterate = 0
            for row in sheet.iter_rows(min_row=_range[0], max_row=_range[1], min_col=_range[2], max_col=_range[3]):
                column_iterate = 0
                if len(matrix) == 1:  # The code is analyzing the chip register matrix
                    if row_iterate == 11 or row_iterate == 12 or row_iterate == 17 or row_iterate == 18:
                        bool_flag = True
                    else:
                        bool_flag = False

                for val in row:
                    if bool_flag:
                        val.value = bool(matrix[chip_iterate][row_iterate][column_iterate])
                    else:
                        val.value = matrix[chip_iterate][row_iterate][column_iterate]
                    column_iterate += 1
                row_iterate += 1
            chip_iterate += 1

    def send_pixel_reg_matrix(self, ranges: List[str], matrix: NDArrayUint32):
        """
        Used for write pixel register table
        :param ranges: Excel cells to get, Ex: ["B2:AE20", ]
        :param matrix: Chip Register matrix to write
        """
        _bool = False
        i = 0
        for sheet_name in self.sheet_names[1:]:  # Getting all sheets except the first one.
            # Add all the sheet positions where dtypes are boolean
            if i == 0 or i == 1 or i == 2 \
                    or i == 4 or i == 5 or i == 6 \
                    or i == 8 or i == 9 or i == 10 \
                    or i == 12 or i == 13 or i == 14 \
                    or i == 16 or i == 17 or i == 18 \
                    or i == 20 or i == 21 or i == 22 \
                    or i == 24 or i == 25 or i == 26 \
                    or i == 28 or i == 29 or i == 30 \
                    or i == 32 or i == 33 or i == 34 \
                    or i == 35 or i == 36 or i == 41 or i == 42 or i == 43:
                _bool = True
            else:
                _bool = False

            self.matrix_to_range_val(ranges, sheet_name, matrix[i], _bool)
            i += 1

